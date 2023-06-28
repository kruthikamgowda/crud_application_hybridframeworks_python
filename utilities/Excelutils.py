import openpyxl
def getRowCount(file,sheetname):
    # load the file(wb)

    wb=openpyxl.load_workbook(file)
    #get the sheet from wb
    sh=wb[sheetname]
    #from the sheet we will get number of rows
    return (sh.max_row)
def getColumnCount(file,sheetname):
    # load the file(wb)

    wb=openpyxl.load_workbook(file)
    #get the sheet from wb
    sh=wb[sheetname]
    #from the sheet we will get number of columns
    return (sh.max_column)

def readdata(file,sheetname,rownumber,columnumber):
    # load the file(wb)

    wb=openpyxl.load_workbook(file)
    #get the sheet from wb
    sh=wb[sheetname]
    #from the sheet we will get data of the cell
    return sh.cell(row=rownumber,column=columnumber).value

def writeData(file,sheetname,rownumber,columnumber,data):
    # load the file(wb)

    wb=openpyxl.load_workbook(file)
    #get the sheet from wb
    sh=wb[sheetname]
    #from the sheet we will get data of the cell
    sh.cell(row=rownumber,column=columnumber).value=data
    wb.save(file)






